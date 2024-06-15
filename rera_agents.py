import concurrent.futures
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException

# Configure Selenium WebDriver
options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")

url = "https://maharera.maharashtra.gov.in/agents-search-result"
service = Service(executable_path="./chromedriver")
driver = webdriver.Chrome(service=service, options=options)
driver.get(url)
links = []  # to store the links
df = pd.read_excel(
    "CP data fields to be scraped (1).xlsx", sheet_name="RERA Agent Data"
)
key_names = df.iloc[:, 0].tolist()
df_scraped = pd.DataFrame(columns=key_names)
certificate_number_tracker = 0


# This function extracts the links of the view details button from each RERA agent
def extract_links(soup):
    global links, certificate_number_tracker
    for row in soup.find("tbody").find_all("tr"):
        cols = row.find_all("td")
        df_scraped.loc[certificate_number_tracker] = {
            "_id": certificate_number_tracker,
            "Professional_Rera_certificate_no": cols[2].text.strip(),
        }
        certificate_number_tracker += 1
        view_details = cols[3].find("a")["href"]
        links.append(view_details)


# This function gets the text of the required field and returns the next element
def find_text(label, soup):
    element = soup.find(string=label)
    if element:
        next_element = element.find_next()
        return next_element.text.strip() if next_element else "N/A"
    return "N/A"


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Adding some padding, max width 50
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_progress(df_scraped, file_path="output.xlsx"):
    df_scraped.to_excel(file_path, index=False)
    adjust_column_width(file_path)


########################################LINKS EXTRACTION STARTS HERE########################################
page_num = 1
while True:
    try:
        if page_num == 4:
            break
        table = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.XPATH, '//div[@class="tableBox"]//div[@class="tableOuter"]//table')
            )
        )
        html_content = table.get_attribute("outerHTML")
        soup = BeautifulSoup(html_content, "html.parser")
        extract_links(soup)
        try:
            soup = BeautifulSoup(driver.page_source, "html.parser")
            next_button = soup.find("a", {"class": "next"})
            if next_button and "disabled" in next_button.get("class", []):
                print("No more pages to explore")
                break
            if next_button:
                next_button_href = next_button["href"]
                driver.get(next_button_href)
                WebDriverWait(driver, 5).until(EC.staleness_of(table))
                page_num += 1
            else:
                print("No next button found")
                break
        except Exception as e:
            print("Error while finding or clicking the next button:", e)
            break
    except Exception as e:
        print("Error during table extraction:", e)
        break

########################################LINK EXTRACTION ENDS########################################


def process_link(link):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(service=Service("./chromedriver"), options=options)
    details = {}
    try:
        driver.get(link)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        soup = BeautifulSoup(driver.page_source, "html.parser")
        details = {
            "Information_Type": find_text("Information Type", soup),
            "First_Name": find_text("First Name", soup),
            "Middle_Name": find_text("Middle Name", soup),
            "Last_Name": find_text("Last Name", soup),
            "Any_criminal_or_police_case_cases_pending": find_text(
                "Any criminal or police case/ cases pending", soup
            ),
            "Father_Full_Name": find_text("Father Full Name", soup),
            "House_Number": find_text("House Number", soup),
            "Building_Name": find_text("Building Name", soup),
            "Street_Name": find_text("Street Name", soup),
            "Locality": find_text("Locality", soup),
            "Landmark": find_text("Land mark", soup),
            "State": find_text("State/UT", soup),
            "Division": find_text("Division", soup),
            "District": find_text("District", soup),
            "Taluka": find_text("Taluka", soup),
            "Village": find_text("Village", soup),
            "PinCode": find_text("Pin Code", soup),
            "Office_Number": find_text("Office Number", soup),
            "Website_URL": find_text("Website URL", soup),
            "ProjectName": find_text("Name", soup),
            "Type Of Project": find_text("Organization Type", soup),
            "Sr.No.": [],
            "Branch_Name": [],
            "LandLine_Number": [],
            "Branch_Address": [],
            "Email_ID": [],
            "Fax_Number": [],
            "Promoter_Name": [],
            "Project_Name": [],
            "Promoted_Certificate_Number": [],
        }
        try:
            details["Do_you_have_any_Past_Experience"] = driver.find_element(
                By.XPATH,
                '//div[contains(text(),"Do you have any Past Experience ?")]/following-sibling::div[1]',
            ).text
        except NoSuchElementException:
            details["Do_you_have_any_Past_Experience"] = "N/A"

        # Below is the past experience table extraction
        try:
            table = WebDriverWait(driver, 1).until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        '//h2[contains(text(),"Past Experience Details")]/following-sibling::div[@id="DataGrid"]//table',
                    )
                )
            )
            soup = BeautifulSoup(table.get_attribute("outerHTML"), "html.parser")
            number_of_past_experience = None
            for row in soup.find("tbody").find_all("tr"):
                if row.find_next().name != "tr":
                    number_of_past_experience = row.find_all("td")[0].text
            if number_of_past_experience:
                details["Past_Experience_Projects_Count"] = number_of_past_experience
        except (TimeoutException, NoSuchElementException):
            details["Past_Experience_Projects_Count"] = "N/A"

        try:
            details["Agent_Registration_in_Other_State"] = driver.find_element(
                By.XPATH,
                '//h2[contains(text(),"Agent Registration in Other State")]/../following-sibling::h4[1]',
            ).get_attribute("textContent")
        except:
            details["Agent_Registration_in_Other_State"] = "N/A"

        try:
            table = WebDriverWait(driver, 1).until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        '//h2[contains(text(),"Branch Details")]/../following-sibling::div[1]//div[1]//div[1]//table',
                    )
                )
            )
            soup = BeautifulSoup(table.get_attribute("outerHTML"), "html.parser")
            for row in soup.find("tbody").find_all("tr"):
                cols = row.find_all("td")
                details["Sr.No."].append(cols[0].text)
                details["Branch_Name"].append(cols[1].text)
                details["LandLine_Number"].append(cols[2].text)
                details["Branch_Address"].append(cols[3].text)
                details["Email_ID"].append(cols[4].text)
                details["Fax_Number"].append(cols[5].text)
                details["Multiple_Branches"] = "Yes"
            details["Sr.No."] = ",".join(details["Sr.No."])
            details["Branch_Name"] = ",".join(details["Branch_Name"])
            details["LandLine_Number"] = ",".join(details["LandLine_Number"])
            details["Branch_Address"] = ",".join(details["Branch_Address"])
            details["Email_ID"] = ",".join(details["Email_ID"])
            details["Fax_Number"] = ",".join(details["Fax_Number"])
        except (TimeoutException, NoSuchElementException):
            details["Sr.No."] = "N/A"
            details["Branch_Name"] = "N/A"
            details["LandLine_Number"] = "N/A"
            details["Branch_Address"] = "N/A"
            details["Email_ID"] = "N/A"
            details["Multiple_Branches"] = "No"

        try:
            table = driver.find_element(
                By.XPATH,
                '//h3[contains(text(),"Promoter Details")]/../../following-sibling::div[1]//div//table',
            )
            soup = BeautifulSoup(table.get_attribute("outerHTML"), "html.parser")
            for row in soup.find("tbody").find_all("tr"):
                cols = row.find_all("td")
                details["Promoter_Name"].append(cols[0].text)
                details["Project_Name"].append(cols[1].text)
                details["Promoted_Certificate_Number"].append(cols[2].text)
            details["Promoter_Name"] = ",".join(details["Promoter_Name"])
            details["Promoted_Projects"] = ",".join(details["Project_Name"])
            details["Project_Name"] = ",".join(details["Project_Name"])
            details["Promoted_Certificate_Number"] = ",".join(
                details["Promoted_Certificate_Number"]
            )
        except NoSuchElementException:
            details["Promoter_Name"] = "N/A"
            details["Project_Name"] = "N/A"
            details["Promoted_Projects"] = "N/A"
            details["Promoted_Certificate_Number"] = "N/A"
        return details
    except Exception as e:
        print(f"Error processing link {link}: {e}")
        return None
    finally:
        driver.quit()


storing = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
    iterable = list(executor.map(process_link, links))
    for it in iterable:
        for key, value in it.items():
            if key == "_id" or key == "Professional_Rera_certificate_no":
                continue
            df_scraped.at[storing, key] = value
        storing += 1
        if (storing % 10 == 0) or (storing == len(links)):
            save_progress(df_scraped)

save_progress(df_scraped)


# Close the main driver
driver.quit()
