import concurrent.futures
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import pickle
import logging
import requests
import traceback

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_progress(df):
    df.to_excel("up.xlsx", index=False)
    adjust_column_width("up.xlsx")


url = "https://up-rera.in/agents"
options = Options()
##options.add_argument("--headless")
options.add_argument("--no-sandbox")

# df = pd.DataFrame(columns=["SrNo.", "Registration ID", "Name", "Address", "Type"])
df = pd.read_excel("/states_data/up.xlsx")
df["Registration Date"] = None

driver = webdriver.Chrome(
    service=Service(executable_path="./chromedriver"), options=options
)
driver.get(url)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
table = driver.find_element(
    By.XPATH,
    "//table[contains(@class,'table') and contains(@class,'table-bordered') and contains(@class,'table-striped')]",
)
counter = 0
try:
    for row in table.find_elements(By.XPATH, "./tbody[1]/tr")[counter:]:
        details = {}
        all_cells = row.find_elements(By.XPATH, "./td")
        view_detail_button = all_cells[5].find_element(
            By.XPATH, "./a[contains(text(),'View Detail')]"
        )
        view_detail_button.click()
        WebDriverWait(driver, 10).until(EC.new_window_is_opened(driver.window_handles))
        driver.switch_to.window(driver.window_handles[-1])
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        registration_date = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'Registration Date :')]/following-sibling::*[1]",
        ).text
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        table = driver.find_element(
            By.XPATH,
            "//table[contains(@class,'table') and contains(@class,'table-bordered') and contains(@class,'table-striped')]",
        )

        # local_driver = webdriver.Chrome(
        #     service=Service(executable_path="./chromedriver"), options=options
        # )

        # details["SrNo."] = all_cells[0].text
        # details["Registration ID"] = all_cells[1].text
        # details["Name"] = all_cells[2].text
        # details["Address"] = all_cells[3].text
        # details["Type"] = all_cells[4].text
        df.loc[counter, "Registration Date"] = details
        counter += 1
        save_progress(df)
        logging.info(f"Scraped and saved {counter+1} data items")
except Exception as e:
    print(e)
    traceback.print_exc
finally:
    driver.quit()
