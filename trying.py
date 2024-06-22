import concurrent.futures
import logging
import re
from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

df = pd.read_excel(
    "CP data fields to be scraped (1).xlsx", sheet_name="RERA Agent Data"
)
key_names = df.iloc[:, 0].tolist()
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def find_text(label, soup):
    element = soup.find(string=label)
    if element:
        next_element = element.find_next()
        return next_element.text.strip() if next_element else "N/A"
    return "N/A"


def save_progress(df_scraped, file_path="output.xlsx"):
    df_scraped.to_excel(file_path, index=False)
    adjust_column_width(file_path)


def each_page(page_num):
    list_of_dictionaries = []
    url = f"https://maharera.maharashtra.gov.in/agents-search-result?agent_name=&agent_project_name=&agent_location=&agent_state=27&agent_division=&agent_district=&page={page_num}&op=Search"
    logging.info(f"Fetching page {page_num}")
    response = requests.get(url)
    if response.status_code != 200:
        logging.error(
            f"Failed to fetch page {page_num}: Status code {response.status_code}"
        )
        return []

    soup = BeautifulSoup(response.content, "html.parser")
    table = soup.find("table")

    if not table:
        logging.warning(f"No table found on page {page_num}")
        return []

    tbody = table.find("tbody")
    if not tbody:
        logging.warning(f"No tbody found in table on page {page_num}")
        return []

    for row_index, row in enumerate(tbody.find_all("tr")):
        details = {}
        data = row.find_all("td")
        if len(data) < 4:
            logging.warning(f"Insufficient data in row {row_index} on page {page_num}")
            continue
        details["_id"] = data[0].text.strip()
        details["Professional_Rera_certificate_no"] = data[2].text.strip()
        link = data[3].find("a")
        if not link or "href" not in link.attrs:
            logging.warning(f"No valid link in row {row_index} on page {page_num}")
            continue

        link_url = link["href"]
        logging.info(f"Fetching details from {link_url}")
        local_response = requests.get(link_url)
        if local_response.status_code != 200:
            logging.error(
                f"Failed to fetch details from {link_url}: Status code {local_response.status_code}"
            )
            continue

        local_soup = BeautifulSoup(local_response.content, "html.parser")
        details.update(
            {
                "Information_Type": find_text("Information Type", local_soup),
                "First_Name": find_text("First Name", local_soup),
                "Middle_Name": find_text("Middle Name", local_soup),
                "Last_Name": find_text("Last Name", local_soup),
                "Any_criminal_or_police_case_cases_pending": find_text(
                    "Any criminal or police case/ cases pending", local_soup
                ),
                "Father_Full_Name": find_text("Father Full Name", local_soup),
                "House_Number": find_text("House Number", local_soup),
                "Building_Name": find_text("Building Name", local_soup),
                "Street_Name": find_text("Street Name", local_soup),
                "Locality": find_text("Locality", local_soup),
                "Landmark": find_text("Land mark", local_soup),
                "State": find_text("State/UT", local_soup),
                "Division": find_text("Division", local_soup),
                "District": find_text("District", local_soup),
                "Taluka": find_text("Taluka", local_soup),
                "Village": find_text("Village", local_soup),
                "PinCode": find_text("Pin Code", local_soup),
                "Office_Number": find_text("Office Number", local_soup),
                "Website_URL": find_text("Website URL", local_soup),
                "ProjectName": find_text("Name", local_soup),
                "Type Of Project": find_text("Organization Type", local_soup),
                "Sr.No.": [],
                "Branch_Name": [],
                "LandLine_Number": [],
                "Branch_Address": [],
                "Email_ID": [],
                "Fax_Number": [],
                "Promoter_Name": [],
                "Project_Name": [],
                "Promoted_Certificate_Number": [],
            }
        )

        experience_element = local_soup.find(
            string=re.compile("Do you have any Past Experience ?", re.IGNORECASE)
        )
        details["Do_you_have_any_Past_Experience"] = (
            experience_element.parent.text.strip() if experience_element else "N/A"
        )

        try:
            # Multiple branches extraction
            branch_details_element = local_soup.find(
                "h2", string=re.compile("Branch Details", re.IGNORECASE)
            )
            if branch_details_element:
                branch_table = branch_details_element.find_next("table")
                if branch_table:
                    branch_tbody = branch_table.find("tbody")
                    if branch_tbody:
                        for branch_row_index, row in enumerate(
                            branch_tbody.find_all("tr")
                        ):
                            cols = row.find_all("td")
                            details["Sr.No."].append(cols[0].text.strip())
                            details["Branch_Name"].append(cols[1].text.strip())
                            details["LandLine_Number"].append(cols[2].text.strip())
                            details["Branch_Address"].append(cols[3].text.strip())
                            details["Email_ID"].append(cols[4].text.strip())
                            details["Fax_Number"].append(cols[5].text.strip())
                        # Joining lists to create a single string for each field
                        details["Sr.No."] = ", ".join(details["Sr.No."])
                        details["Branch_Name"] = ", ".join(details["Branch_Name"])
                        details["LandLine_Number"] = ", ".join(
                            details["LandLine_Number"]
                        )
                        details["Branch_Address"] = ", ".join(details["Branch_Address"])
                        details["Email_ID"] = ", ".join(details["Email_ID"])
                        details["Fax_Number"] = ", ".join(details["Fax_Number"])
                    else:
                        logging.warning(
                            f"No tbody found in branch details table on page {page_num}"
                        )
                        details["Sr.No."] = "N/A"
                        details["Branch_Name"] = "N/A"
                        details["LandLine_Number"] = "N/A"
                        details["Branch_Address"] = "N/A"
                        details["Email_ID"] = "N/A"
                        details["Fax_Number"] = "N/A"
        except Exception as e:
            logging.error(f"Error extracting branch details on page {page_num}: {e}")
            details["Sr.No."] = "N/A"
            details["Branch_Name"] = "N/A"
            details["LandLine_Number"] = "N/A"
            details["Branch_Address"] = "N/A"
            details["Email_ID"] = "N/A"
            details["Fax_Number"] = "N/A"

        try:
            # Promoter details extraction
            promoter_details_element = local_soup.find(
                "h3", string=re.compile("Promoter Details", re.IGNORECASE)
            )
            if promoter_details_element:
                promoter_table = promoter_details_element.find_next("table")
                if promoter_table:
                    promoter_tbody = promoter_table.find("tbody")
                    if promoter_tbody:
                        for promoter_row_index, row in enumerate(
                            promoter_tbody.find_all("tr")
                        ):
                            cols = row.find_all("td")
                            details["Promoter_Name"].append(cols[0].text.strip())
                            details["Project_Name"].append(cols[1].text.strip())
                            details["Promoted_Certificate_Number"].append(
                                cols[2].text.strip()
                            )
                        details["Promoter_Name"] = ", ".join(details["Promoter_Name"])
                        details["Project_Name"] = ", ".join(details["Project_Name"])
                        details["Promoted_Certificate_Number"] = ", ".join(
                            details["Promoted_Certificate_Number"]
                        )
                        details["Promoted_Projects"] = ",".join(details["Project_Name"])
                    else:
                        logging.warning(
                            f"No tbody found in promoter details table on page {page_num}"
                        )
                        details["Promoter_Name"] = "N/A"
                        details["Project_Name"] = "N/A"
                        details["Promoted_Certificate_Number"] = "N/A"
                        details["Promoted_Projects"] = "N/A"
        except Exception as e:
            logging.error(f"Error extracting promoter details on page {page_num}: {e}")
            details["Promoter_Name"] = "N/A"
            details["Project_Name"] = "N/A"
            details["Promoted_Certificate_Number"] = "N/A"
            details["Promoted_Projects"] = "N/A"

        list_of_dictionaries.append(details)
        logging.info(f"Processed row {row_index} on page {page_num}")
    logging.info(f"Completed fetching page {page_num}")
    return list_of_dictionaries


def scrape_pages(page_range):
    all_details = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        results = executor.map(each_page, page_range)
        for result in results:
            all_details.extend(result)
    return all_details


# Scrape pages and collect data
page_range = range(1, 4733)
all_details = scrape_pages(page_range)

# Save collected data
data_frame = pd.DataFrame(all_details)
save_progress(data_frame)
