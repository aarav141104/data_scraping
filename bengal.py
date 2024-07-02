import concurrent.futures
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import pickle
import logging
import requests

url = "https://rera.wb.gov.in/agent_list.php?dcode=0"
options = Options()
# options.add_argument("--headless")
options.add_argument("--no-sandbox")
driver = webdriver.Chrome(
    service=Service(executable_path="./chromedriver"), options=options
)
driver.get(url)
WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))


df = pd.DataFrame(
    columns=[
        "Sl No.",
        "Agent ID",
        "Agent Name",
        "Registered From",
        "Registration Number",
        "Registration Date",
    ]
)
counter = 0


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_progress(df):
    df.to_excel("bengal_agents.xlsx", index=False)
    adjust_column_width("bengal_agents.xlsx")


while True:
    try:
        # response = requests.get(url)
        # soup = BeautifulSoup(response.content, "html.parser")
        table = driver.find_element(By.TAG_NAME, "table")
        if table == None:
            print("lag gaye")
        rows = table.find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr")
        for row in rows:
            dict_scraped = {}
            data = row.find_elements(By.TAG_NAME, "td")
            print(len(data))
            dict_scraped["Sl No."] = data[0].text.strip()
            dict_scraped["Agent ID"] = data[1].text.strip()
            dict_scraped["Agent Name"] = data[2].text.strip()
            dict_scraped["Registered From"] = data[3].text.strip()
            dict_scraped["Registration Number"] = data[4].text.strip()
            dict_scraped["Registration Date"] = data[5].text.strip()
            df.loc[counter] = dict_scraped
            counter += 1
        next_element = driver.find_element(By.LINK_TEXT, "Next")
        if next_element.get_attribute("class").split()[-1] == "disabled":
            save_progress(df)
            break
        next_element.click()
        # soup = BeautifulSoup(driver.page_source, "html.parser")
    except Exception as e:
        print(e)
