from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import pdfkit
import os
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import concurrent.futures
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options

service = Service(executable_path="./chromedriver")
driver = webdriver.Chrome(service=service)
url = "https://www.magicbricks.com/residential-real-estate-agents-in-mumbai-pppagent"
xpath = "//span[contains(@class,'seeProDetail')]//a"
df = pd.read_excel("CP data fields to be scraped (1).xlsx", sheet_name="Portals Data")
magicbricks_fields = df.iloc[:, 0].dropna().tolist()  # Fields for Magicbricks
df_magic = pd.DataFrame(columns=magicbricks_fields)
magic_dict = {field: None for field in magicbricks_fields}
driver.get(url)
see_details = WebDriverWait(driver, 6).until(
    EC.presence_of_all_elements_located((By.XPATH, xpath))
)


def save_progress(df_scraped, file_path="output.xlsx"):
    df_scraped.to_excel(file_path, index=False)
    adjust_column_width(file_path)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Adding some padding, max width 50
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def properties_for_sale_1(url):
    global magic_dict
    driver = webdriver.Chrome(service=Service("./chromedriver"))
    driver.get(url)
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )
    try:
        # Locate the "See all" button
        see_all_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    '//a[contains(@class,"prop_sale_seeAll") and contains(text(),"Residential")]',
                )
            )
        )
        see_all_button.click()
        driver.implicitly_wait(10)
        new_tab = driver.window_handles[-1]
        driver.switch_to.window(new_tab)
        WebDriverWait(driver, 12).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        all_elements = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, '//div[contains(@class,"mb-srp__list")]')
            )
        )
        for element in all_elements:
            element.click()
            driver.implicitly_wait(5)
            new_tab_2 = driver.window_handles[-1]
            driver.switch_to.window(new_tab_2)
            print(driver.current_url)
    except:
        print("something")
    finally:
        print("finally")
        # try:
        #     project = element.find_element(
        #         By.XPATH, './/h2[contains(@class,"mb-srp__card--title")]'
        #     )
        #     driver.execute_script("arguments[0].click();", project)
        #     print("Clicked on project title.")

        #     WebDriverWait(driver, 10).until(
        #         EC.presence_of_element_located((By.TAG_NAME, "body"))
        #     )

        #     magic_dict["Properties for sale"] = (
        #         "Ticket Size : "
        #         + driver.find_element(
        #             By.XPATH, '//div[contains(@class,"mb-ldp__dtls__price")]'
        #         ).get_attribute("textContent")
        #     )
        #     break
        # except NoSuchElementException:
        #     print(
        #         "Project title or price element not found within the property element."
    #             #)
    # except NoSuchElementException as e:
    #     print(f"An error occurred: NoSuchElementException: {e}")
    # except TimeoutException as e:
    #     print(f"An error occurred: TimeoutException: {e}")
    # except Exception as e:
    #     print(f"An unexpected error occurred: {e}")


def process_link(see_detail):
    global magic_dict
    detail_url = see_detail.get_attribute("href")
    driver.get(detail_url)
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )
    try:
        magic_dict["Name"] = driver.find_element(
            By.XPATH,
            '//div[@class="fedImg"]/following-sibling::span[@class="agntName"]',
        ).text
    except:
        magic_dict["Name"] = "N/A"

    try:
        magic_dict["Company Name"] = driver.find_element(
            By.XPATH, '//div[@class="agentNameLoc"]//div[@class="agentName"]'
        ).text
    except:
        magic_dict["Company Name"] = "N/A"

    magic_dict["RERA ID "] = detail_url.split("-")[-1]

    try:
        magic_dict["Operating since"] = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Operating Since")]/following-sibling::div[1]',
        ).text
    except:
        magic_dict["Operating since"] = "N/A"

    try:
        first = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Properties for Sale") and contains(@class,"column_1")]/following-sibling::div[1]',
        )
        properties_for_sale = first.text
        magic_dict["Properties For Sale"] = properties_for_sale
    except:
        magic_dict["Properties For Sale"] = "N/A"

    try:
        first = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Properties for Rent") and contains(@class,"column_1")]/following-sibling::div[1]',
        )
        properties_for_rent = first.text
        magic_dict["Properties For rent"] = properties_for_rent
    except:
        magic_dict["Properties For rent"] = "N/A"

    try:
        first = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Address") and contains(@class,"column_1")]/following-sibling::div[1]',
        )
        second = first.find_element(
            By.XPATH, "./following-sibling::br[1]/following-sibling::text()"
        )
        magic_dict["Address"] = first.text + "\n" + second
    except:
        magic_dict["Address"] = "N/A"

    try:
        magic_dict["Deals in"] = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Dealing In") and contains(@class,"column_1")]/following-sibling::div[1]',
        ).text
    except:
        magic_dict["Deals in"] = "N/A"

    try:
        more_button = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Operating In") and contains(@class,"column_1")]/following-sibling::div[1]//span[1]//a[contains(text(),"+ more")]',
        )
        more_button.click()
        operates_in = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"Operating In") and contains(@class,"column_1")]/following-sibling::div[1]//span[2]',
        )
        operates_in_data = ""
        for item in operates_in.find_elements(By.XPATH, ".//a"):
            # if item.find_element(By.XPATH, "./following-sibling::*[1]").tag_name != "a":
            #     break
            operates_in_data += item.text + ","
        magic_dict["Operates in"] = operates_in_data
    except:
        magic_dict["Operates in"] = "N/A"

    try:
        first = driver.find_element(
            By.XPATH,
            '//div[contains(text(),"About the Agent")]/following-sibling::div[1]//span[1]',
        ).text

        try:
            more_button = driver.find_element(
                By.XPATH,
                '//div[contains(@class,"highlightsInfo aboutAgentTxt")]//a[contains(@class,"moreData") and contains(text(),"+ more")]',
            )
            driver.execute_script("arguments[0].click();", more_button)

            try:
                second = driver.find_element(
                    By.XPATH,
                    '//div[contains(text(),"About the Agent")]/following-sibling::div[1]//span[2]',
                ).text
            except NoSuchElementException:
                print("Second span not found")
                second = ""

        except NoSuchElementException:
            print("More button not found")
            second = ""

        magic_dict["About Company"] = first + second

    except NoSuchElementException:
        print("Some issue with locating elements")
        magic_dict["About Company"] = "N/A"
    properties_for_sale_1(detail_url)


c = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
    iterable = list(executor.map(process_link, see_details[:10]))
    for it in iterable:
        df_magic.loc[c] = it
    c += 1
    if (c % 10 == 0) or (c == len(see_details)):
        save_progress(df_magic)

save_progress(df_magic)


driver.quit()
