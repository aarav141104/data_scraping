import concurrent.futures
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import pickle
import logging
import requests
import traceback

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)


def save_progress(df):
    df.to_excel("magicbricks.xlsx", index=False)
    adjust_column_width("magicbricks.xlsx")


options = Options()
# options.add_argument("--headless")
options.add_argument("--no-sandbox")

url = "https://www.magicbricks.com/residential-real-estate-agents-in-mumbai-pppagent"
driver = webdriver.Chrome(
    service=Service(executable_path="./chromedriver"), options=options
)
driver.get(url)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

df = pd.DataFrame(
    columns=[
        "About Company",
        "Deals in",
        "Company Name",
        "RERA ID",
        "Name",
        "Operating since",
        "Properties For Sale",
        "Properties For Rent",
        "Address",
        "Operates In",
        "Project",
        "Ticket Size",
        "Location",
        "Config",
    ]
)
page_num = 1
link_in_a_page = 1
property_number = 1

visited_urls = set()
counter = 0
while True:
    try:
        logging.info(f"in page number {page_num}")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        elements = driver.find_elements(
            By.XPATH, "//span[contains(@class,'seeProDetail')]/a[1]"
        )
        urls_of_each_page = [element.get_attribute("href") for element in elements]
        print(f"amount of links in this page : {len(urls_of_each_page)}")
        for url in urls_of_each_page:
            logging.info(f"IN LINK NUMBER {link_in_a_page}")
            details = {}
            local_driver = webdriver.Chrome(
                service=Service(executable_path="./chromedriver"), options=options
            )
            if url not in visited_urls:
                local_driver.get(url)
            else:
                continue
            WebDriverWait(local_driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            ##FUCK THESE LINES BELOW
            # try:
            #     more_data_buttons = local_driver.find_elements(By.CLASS_NAME, "moreData")
            # except NoSuchElementException:
            #     print("Error while extracting the more buttons")
            # if "showFullAboutVeriAgent();" in more_data_buttons[0].get_attribute(
            #     "onclick"
            # ):
            #     more_data_buttons[0].click()
            #####TILL HERE
            first_case = None
            try:
                more_button = local_driver.find_element(
                    By.XPATH,
                    "//div[contains(text(),'About the Agent')]/following-sibling::div[1]/a[1]",
                )
                try:
                    first_case = local_driver.find_element(
                        By.XPATH, '//span[contains(@id,"shortDescVre")]'
                    ).text
                    first_case += local_driver.find_element(
                        By.XPATH, '//span[contains(@id,"fullDescvery")]'
                    ).text
                    details["About Company"] = first_case
                except NoSuchElementException:
                    pass
                if not first_case:
                    try:
                        second_case = local_driver.find_element(
                            By.XPATH, "//span[contains(@id,'fullDesc')]"
                        ).text
                        details["About Company"] = second_case
                    except NoSuchElementException:
                        pass
            except NoSuchElementException as e:
                details["About Company"] = local_driver.find_element(
                    By.XPATH,
                    "//div[contains(text(),'About the Agent')]/following-sibling::div[1]/span[1]",
                ).text
            try:
                details["Deals in"] = local_driver.find_element(
                    By.XPATH,
                    '//div[contains(text(),"Dealing In")]/following-sibling::div[1]',
                ).text
            except NoSuchElementException:
                print("some issue with extracting deals in")
            try:
                details["Company Name"] = local_driver.find_element(
                    By.CLASS_NAME, "agentName"
                ).text
            except NoSuchElementException:
                details["Company Name"] = "N/A"
            try:
                details["RERA ID"] = local_driver.current_url.split("-")[-1]
            except IndexError:
                details["RERA ID"] = "N/A"
            try:
                details["Name"] = local_driver.find_element(
                    By.CLASS_NAME, "agntName"
                ).text
            except NoSuchElementException:
                details["Name"] = "N/A"
            try:
                details["Operating since"] = local_driver.find_element(
                    By.XPATH,
                    '//div[contains(text(),"Operating Since")]/following-sibling::div[1]',
                ).text
            except NoSuchElementException:
                details["Operating since"] = "N/A"
            try:
                details["Properties For Sale"] = local_driver.find_element(
                    By.XPATH,
                    "//div[contains(text(),'Properties for Sale')]/following-sibling::div[1]",
                ).text
            except NoSuchElementException:
                details["Properties For Sale"] = "N/A"
            try:
                details["Properties For Rent"] = local_driver.find_element(
                    By.XPATH,
                    "//div[contains(text(),'Properties for Rent')]/following-sibling::div[1]",
                ).text
            except NoSuchElementException:
                details["Properties For Rent"] = "N/A"
            try:
                details["Operates In"] = local_driver.find_element(
                    By.XPATH,
                    "//div[contains(text(),'Operating In')]/following-sibling::div[1]/span[contains(@id,'locFull')][1]",
                ).text
            except NoSuchElementException:
                details["Operates In"] = "N/A"
            properties_button = local_driver.find_element(
                By.XPATH, "//a[contains(@class,'prop_sale_seeAll')]"
            )
            properties_button.click()
            #####The property thing starts from here##############
            the_property_page = local_driver.window_handles[-1]
            local_driver.switch_to.window(the_property_page)
            WebDriverWait(local_driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            local_driver.execute_script(
                "window.scrollTo(0,document.body.scrollHeight);"
            )
            local_driver.implicitly_wait(5)
            all_properties = local_driver.find_elements(By.CLASS_NAME, "mb-srp__list")
            amount_of_properties = len(all_properties)
            ######Essentially ends here#################
            details["Project"] = []
            details["Ticket Size"] = []
            details["Location"] = []
            details["Config"] = []
            print(f"Amount of properties : {amount_of_properties}")
            for i in range(amount_of_properties):
                all_properties[i].click()
                local_driver.switch_to.window(local_driver.window_handles[-1])
                WebDriverWait(local_driver, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                logging.info(f"IN PROPERTY NUMBER {property_number}")
                try:
                    project_name = local_driver.find_element(
                        By.XPATH,
                        "//div[contains(text(),'Project') and contains(@class,'mb-ldp__dtls__body__list--label')]/following-sibling::div[1]",
                    ).text
                except Exception as e:
                    project_name = "N/A"
                details["Project"].append(project_name)
                try:
                    ticket_size = local_driver.find_element(
                        By.XPATH,
                        "//div[contains(@class,'mb-ldp__dtls__flex-row pad-b-4')]/div[contains(@class,'mb-ldp__dtls__price')]",
                    ).text
                except Exception as e:
                    ticket_size = "N/A"
                details["Ticket Size"].append(ticket_size)
                try:
                    location = local_driver.find_element(
                        By.XPATH,
                        "//div[contains(@class,'mb-ldp__more-dtl__list--label') and contains(text(),'Address')]/following-sibling::div[1]",
                    ).text
                except Exception as e:
                    location = "N/A"
                details["Location"].append(location)
                try:
                    config = local_driver.find_element(
                        By.XPATH,
                        "//div[contains(@class,'mb-ldp__dtls__body__summary--left mb-ldp__dtls__body__summary--dflex')]",
                    ).text
                except Exception as e:
                    config = "N/A"
                details["Config"].append(config)
                local_driver.close()
                local_driver.switch_to.window(the_property_page)
                WebDriverWait(local_driver, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                all_properties = local_driver.find_elements(
                    By.CLASS_NAME, "mb-srp__list"
                )
                property_number += 1
            details["Project"] = ",".join(details["Project"])
            details["Ticket Size"] = ",".join(details["Ticket Size"])
            details["Location"] = ",".join(details["Location"])
            details["Config"] = ",".join(details["Config"])
            df.loc[counter] = details
            counter += 1
            local_driver.quit()
            save_progress(df)
            link_in_a_page += 1
            property_number = 1
        next_page = driver.find_element(By.XPATH, "//*[contains(text(),'Next Page')]")
        next_page.click()
        page_num += 1
        link_in_a_page = 1
    except Exception as e:
        print(e)
        traceback.print_exc()
